import pandas as pd
from typing import List, Dict
from models.account import Account
from openpyxl.styles import Font, Alignment

def generate_excel_report(
    summary_data: Dict,
    accounts: List[Account],
    inquiries: List[Dict],
    credit_repair_items: List[Dict],
    output_path: str = "credit_report_summary.xlsx"
):
    """
    Generate an Excel report with a summary tab and detailed data tabs.
    """
    try:
        with pd.ExcelWriter(output_path, engine='openpyxl') as writer:
            # --- Summary Tab ---
            worksheet = writer.book.create_sheet("Summary", 0)
            writer.sheets["Summary"] = worksheet
            
            # Payoff Summary Table
            payoff_df = pd.DataFrame.from_dict(
                summary_data.get("payoff_summary", {}), 
                orient='index', 
                columns=['Amount']
            )
            payoff_df.index.name = "Action"
            payoff_df.to_excel(writer, sheet_name='Summary', startrow=1, startcol=0)
            worksheet.cell(row=1, column=1, value="Pay-Off Goals").font = Font(bold=True)

            # Counts Summary Table
            counts_df = pd.DataFrame.from_dict(
                summary_data.get("counts", {}),
                orient='index',
                columns=['Count']
            )
            counts_df.index.name = "Metric"
            counts_df.to_excel(writer, sheet_name='Summary', startrow=1, startcol=4)
            worksheet.cell(row=1, column=5, value="Report Metrics").font = Font(bold=True)

            # AI Analysis
            ai_analysis = summary_data.get("ai_analysis", "No analysis available.")
            analysis_start_row = max(payoff_df.shape[0], counts_df.shape[0]) + 4
            worksheet.cell(row=analysis_start_row, column=1, value="AI Credit Potential Analysis").font = Font(bold=True)
            cell = worksheet.cell(row=analysis_start_row + 1, column=1, value=ai_analysis)
            cell.alignment = Alignment(wrap_text=True, vertical='top')
            worksheet.column_dimensions['A'].width = 100 # Widen column for analysis

            # --- Data Tabs ---
            # Accounts Tab
            account_dicts = [acc.to_dict() for acc in accounts]
            accounts_df = pd.DataFrame(account_dicts)
            if accounts_df.empty:
                accounts_df = pd.DataFrame(columns=[
                    "Bank", "Account Type", "Open Date", "Responsibility", "Balance", "Limit",
                    "Utilization", "Rating", "A Rating Limit (10%)", "B Rating Limit (20%)",
                    "C Rating Limit (30%)", "Status"
                ])
            accounts_df.to_excel(writer, sheet_name='Accounts', index=False)

            # Inquiries Tab
            inquiries_df = pd.DataFrame(inquiries)
            if inquiries_df.empty:
                inquiries_df = pd.DataFrame(columns=["date", "creditor", "type", "reason"])
            inquiries_df.to_excel(writer, sheet_name='Inquiries', index=False)

            # Credit Repair Tab
            repair_df = pd.DataFrame(credit_repair_items)
            expected_columns = ['BUREAU', 'TYPE', 'Account', 'Occurrence', 'LAST DLQ', 'NOTES', 'INITIAL']
            if repair_df.empty:
                repair_df = pd.DataFrame(columns=expected_columns)
            else:
                # Ensure all expected columns exist
                for col in expected_columns:
                    if col not in repair_df.columns:
                        repair_df[col] = None
                repair_df = repair_df[expected_columns] # Enforce column order
            repair_df.to_excel(writer, sheet_name='Credit Repair', index=False)

        print(f"Successfully generated Excel report at {output_path}")
        return output_path
    except Exception as e:
        print(f"Failed to generate Excel report: {e}")
        return None 